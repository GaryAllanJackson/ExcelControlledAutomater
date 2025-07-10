import os
from contextlib import nullcontext
from datetime import datetime
import time
from urllib.parse import urlparse, parse_qs

import pyautogui
import requests
from selenium.common import WebDriverException
from selenium.webdriver import ActionChains
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
import openpyxl
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

from common import variables, command_library
import pandas
# import pyodbc
from sqlalchemy import create_engine, text

import sys
import contextlib

from common.web_scraper import WebScraper


@contextlib.contextmanager
def suppress_stderr():
    with open(os.devnull, 'w') as f_null:
        stderr = sys.stderr
        sys.stderr = f_null
        try:
            yield
        finally:
            sys.stderr = stderr

class Functions:
    def __init__(self, driver):
        self.driver = driver
        self.log_file_name = variables.log_file_name
        self.log_sheet_name = variables.log_sheet_name
        self.saved_text = ""
        self.command_file = variables.command_file_name
        self.command_sheet = variables.command_sheet_name
        self.data = []
        self.har_file_name = ""
        self.har_file_contents = ""
        self.has_save_complete = False
        self.engine = None
        self.wps = WebScraper()
        self.check_response_url = ""



    def navigate(self, page_url, description):
        if variables.unsecure_protocol in self.driver.current_url or variables.secure_protocol in self.driver.current_url:
            self.get_har_file_for_tag_information()
        self.driver.get(page_url)
        time.sleep(3)
        self.log_equal_action("navigate", page_url, self.driver.current_url, description)
        print(f"page_url: {page_url}")
        print(f"Current_url: {self.driver.current_url}")
        assert page_url == self.driver.current_url, "Navigation Failed!"

    def navigate_without_checking(self, page_url, description):
        print(f"{variables.terminal_color_blue} Navigating to {page_url}{variables.terminal_color_reset}")
        if variables.unsecure_protocol in self.driver.current_url or variables.secure_protocol in self.driver.current_url:
            try:
                self.get_har_file_for_tag_information()
            except:
                print(f"No har information for {self.driver.current_url}")
        self.driver.get(page_url)
        time.sleep(3)
        self.log_equal_action("navigate - no check", page_url, self.driver.current_url, description)

    def perform_click(self, accessor_type, accessor, expected, actual, description):
        element = self.get_element(accessor_type, accessor)
        element.click()
        print(f"expected = {expected}")
        if expected is not None and expected != "n/a":
            time.sleep(3)
            if actual.lower() == variables.get_current_url_in_action_method or actual == "":
                actual = self.driver.current_url
        print(f"In perform_click: Expected = {expected} - Actual = {actual}")
        self.log_equal_action("click", expected, actual, description)
        assert expected == actual, "Click Failed!"

    def perform_print(self, phrase, description):
        if phrase.startswith("=="):
            phrase = "'" + phrase
        if variables.get_current_url_in_action_method in phrase.lower():
            if phrase == variables.get_current_url_in_action_method or phrase == "current url":
                phrase = self.driver.current_url
            elif variables.get_current_url_in_action_method in phrase:
                phrase = phrase.replace(variables.get_current_url_in_action_method, self.driver.current_url)
            else:
                index = phrase.lower().find(variables.get_current_url_in_action_method)
                phrase_len = len(variables.get_current_url_in_action_method)
                phrase = phrase[:index] + self.driver.current_url + phrase[phrase_len:]
        self.log_equal_action("Print", phrase,phrase, description)
        print(phrase)

    def perform_print_sub_elements(self, accessor_type, accessor, description):
        elements = self.get_elements(accessor_type, accessor)
        # elements_have_text = False
        print("-" * variables.dash_length)
        print(f"Printing all child elements based on: {accessor_type} = {accessor}")
        for element in elements:
            elements_have_text = False
            child_elements = element.find_elements(By.CSS_SELECTOR, "*")
            for child_element in child_elements:
                tag_name = child_element.tag_name
                tag_text = self.get_element_text_silently(child_element)
                if len(tag_text) > 0:
                    print(f'Tag Name:{tag_name}: {tag_text}')
                    elements_have_text = True
            if elements_have_text:
                print("-"*variables.dash_length)
        self.log_equal_action("Print Child Elements", "n/a", "n/a", description)

    def perform_print_all_elements(self, accessor_type, accessor, description):
        elements = self.get_elements(accessor_type, accessor)
        elements_have_text = False
        print("-" * variables.dash_length)
        print(f"Printing all elements based on: {accessor_type} = {accessor}")
        for element in elements:
            tag_name = element.tag_name
            tag_text = self.get_element_text_silently(element)
            if len(tag_text) > 0:
                print(f'Tag Name:{tag_name}: {tag_text}')
                elements_have_text = True
        if elements_have_text:
            print("-" * variables.dash_length)
        self.log_equal_action("Print All Elements", "n/a", "n/a", description)

    def perform_select_all_elements(self, accessor_type, accessor, description):
        elements = self.get_elements(accessor_type, accessor)
        elements_have_text = False
        print("-" * variables.dash_length)
        print(f"Printing all elements based on: {accessor_type} = {accessor}")
        for element in elements:
            tag_name = element.tag_name
            tag_text = self.get_element_text_silently(element)
            element.click()
            if len(tag_text) > 0:
                print(f'Tag Name:{tag_name}: {tag_text}')
                elements_have_text = True
        if elements_have_text:
            print("-" * variables.dash_length)
        self.log_equal_action("Select All Elements", "n/a", "n/a", description)

    def perform_page_refresh(self):
        self.driver.refresh()
        time.sleep(3)

    def perform_send_key(self, accessor_type, accessor, send_text, expected, actual, description):
        element = self.get_element(accessor_type, accessor)
        element.send_keys(send_text)
        #in case the actual value needs to be retrieved from the element where text was sent
        print(f"actual (get_text) = {actual}")
        if actual == "get_text":
            actual = self.get_element_text_silently(element)
        elif actual == variables.get_current_url_in_action_method:
            actual = self.driver.current_url
        self.log_equal_action("Send Keys", expected, actual, description)
        assert expected == actual, "Click Failed!"

    def get_element(self, selector_type, selector):
        # print(f"in get_element selector_type = {selector_type}")
        if selector_type.lower() == "xpath":
            return self.driver.find_element(By.XPATH, selector)
        elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
            return self.driver.find_element(By.CSS_SELECTOR, selector)
        elif selector_type.lower() == "id":
            return self.driver.find_element(By.ID, selector)
        elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
            return self.driver.find_element(By.TAG_NAME, selector)
        elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
            return self.driver.find_element(By.LINK_TEXT, selector)
        elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
            return self.driver.find_element(By.CLASS_NAME, selector)
        else:
            print(f"in get_element else statement: selector_type = {selector_type}")
            return self.driver.find_element(By.NAME, selector)

    def get_elements(self, selector_type, selector):
        # print(f"in get_element selector_type = {selector_type}")
        if selector_type.lower() == "xpath":
            return self.driver.find_elements(By.XPATH, selector)
        elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
            return self.driver.find_elements(By.CSS_SELECTOR, selector)
        elif selector_type.lower() == "id":
            return self.driver.find_elements(By.ID, selector)
        elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
            return self.driver.find_elements(By.TAG_NAME, selector)
        elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
            return self.driver.find_elements(By.LINK_TEXT, selector)
        elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
            return self.driver.find_elements(By.CLASS_NAME, selector)
        else:
            print(f"in get_element else statement: selector_type = {selector_type}")
            return self.driver.find_elements(By.NAME, selector)

    def get_element_text(self, accessor_type, accessor, expected, description):
        element = self.get_element(accessor_type, accessor)
        el_html = element.get_attribute("outerHTML")
        if el_html.find("<input") > -1:
            return_value = element.get_attribute("value")
        else:
            try:
                return_value = element.text
            except ValueError:
                return_value = element.get_attribute("innerText")

        actual = return_value if return_value is not None else ""
        print(f"Get Text - Expected:{expected} - Actual:{actual}")
        self.log_equal_action("Get Text", expected, actual, description)
        # Save the value into saved_text unless this method was
        # called to compare this with the saved_text
        if expected != self.saved_text:
            self.saved_text = return_value
            print(f"Retrieved and stored Text: {return_value}")
        else:
            print(f"Retrieved Text: {return_value}")
        return return_value

    def get_element_text_alt(self, element, expected, actual, description):
        el_html = element.get_attribute("outerHTML")
        # returnValue = ""
        if el_html.find("<input") > -1:
            return_value = element.get_attribute("value")
        else:
            try:
                return_value = element.text
            except ValueError as e:
                return_value = element.get_attribute("innerText")
                print(f"Getting innerText {e}")
        print(f"Retrieved Text: {return_value}")
        self.log_equal_action("Get Text", expected, actual, description)
        return return_value if return_value is not None else ""


    @staticmethod
    def get_element_text_silently(element):
        el_html = element.get_attribute("outerHTML")
        # print(f"el_html = {el_html}")
        # return_value = ""
        if el_html.find("<input") > -1:
            return_value = element.get_attribute("value")
        else:
            try:
                return_value = element.text
            except ValueError as e:
                return_value = element.get_attribute("innerText")
                print(f"Getting innerText {e}")

        return return_value if return_value is not None else ""

    def compare_element_text_to_saved_text(self, accessor_type, accessor, actual, expected, description):
        self.get_element_text(accessor_type, accessor, self.saved_text, "Comparing against saved text.")
        print(f"Compared Text: {expected} vs {self.saved_text}")
        self.log_equal_action("Compare Text", expected, actual, description)

    def perform_hover(self, accessor_type, accessor, description):
        element = self.get_element(accessor_type, accessor)
        actions = ActionChains(self.driver)
        actions.move_to_element(element).perform()
        self.log_equal_action("Hover", "n/a", "n/a", description)
        self.perform_print("Hovering over element", description)

    def select_dropdown_by_value(self, accessor_type, accessor, value, description):
        dropdown = Select(self.get_element(accessor_type, accessor))
        dropdown.select_by_value(value)
        self.log_equal_action("Select Dropdown by value", "n/a", "n/a", description)


    def perform_screenshot(self, screenshot_file_name, description):
        try:
            screenshot = pyautogui.screenshot()
            screenshot.save("screenshots/" + screenshot_file_name)
            file_saved = True
        except Exception as e:
            print(e)
            self.driver.save_screenshot("screenshots/" + screenshot_file_name)
            file_saved = True
        file_saved = file_saved if file_saved is not None else False
        self.log_equal_action("Take Screenshot", "True", str(file_saved),description)

    def perform_wait(self, text_url, description):
        time.sleep(int(text_url))
        self.log_equal_action("Wait", "n/a", "n/a", description)

    def presence_of_element_located(self, selector_type, selector, wait_time, description):
        # wait = WebDriverWait(self.driver, 20)
        wait = WebDriverWait(self.driver, int(wait_time))
        if selector_type.lower() == "xpath":
            wait.until(expected_conditions.presence_of_element_located((By.XPATH, selector)))
        elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
            wait.until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, selector)))
        elif selector_type.lower() == "id":
            wait.until(expected_conditions.presence_of_element_located((By.ID, selector)))
        elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
            wait.until(expected_conditions.presence_of_element_located((By.TAG_NAME, selector)))
        elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
            wait.until(expected_conditions.presence_of_element_located((By.LINK_TEXT, selector)))
        elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
            wait.until(expected_conditions.presence_of_element_located((By.CLASS_NAME, selector)))
        else:
            wait.until(expected_conditions.presence_of_element_located((By.NAME, selector)))
        element = self.get_element(selector_type, selector)
        element_found = element is not None
        self.log_equal_action("Wait for Element Presence", "True", str(element_found), description)

    def switch_to_window(self, index, description):
        win_handle = self.driver.window_handles
        self.driver.switch_to.window(win_handle[int(index)])
        self.log_equal_action("Switch to Window", "n/a", "n/a", description)

    # Waits for an element to be clickable and returns that element to the calling method
    # placed in a try catch block to avoid returning the element if the wait expires before
    # the element is clickable
    def wait_for_element_to_be_clickable(self, selector_type, selector, wait_time, description):
        # wait = WebDriverWait(driver, 20)
        print(f"Waiting {wait_time} seconds for element to be clickable.")
        if wait_time is None:
            wait_time = 20
        wait = WebDriverWait(self.driver, int(wait_time))
        try:
            if selector_type.lower() == "xpath":
                wait.until(expected_conditions.element_to_be_clickable((By.XPATH, selector)))
            elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
                wait.until(expected_conditions.element_to_be_clickable((By.CSS_SELECTOR, selector)))
            elif selector_type.lower() == "id":
                wait.until(expected_conditions.element_to_be_clickable((By.ID, selector)))
            elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
                wait.until(expected_conditions.element_to_be_clickable((By.TAG_NAME, selector)))
            elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
                wait.until(expected_conditions.element_to_be_clickable((By.LINK_TEXT, selector)))
            elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
                wait.until(expected_conditions.element_to_be_clickable((By.CLASS_NAME, selector)))
            else:
                wait.until(expected_conditions.element_to_be_clickable((By.NAME, selector)))
            element = self.get_element(selector_type, selector)
            self.log_equal_action("Wait for clickable element", "n/a", "n/a", description)
        except ValueError as e:
            element = None
            self.log_equal_action("Wait for clickable element", "n/a", "None", description + " - " + e)
        return element

    def get_json_from_api(self, api_url, description):
        response = requests.get(api_url)
        if response.status_code == 200:
            # json_data = requests.get(api_url).json()
            print(f"Successful API connection. Status code: {response.status_code}")
            json_data = response.json()
            json_len = len(json_data)
            field_lens_printed = False
            print(f"JSON API - Number of records returned: {json_len}")
            for i, item in enumerate(json_data, start=1):
                if not field_lens_printed:
                    print(f"JSON API - Number of fields per record: {len(item)}")
                    field_lens_printed = True
                fields_output = "{"
                for key, value in item.items():
                    print(f"{i} - {key}:{value}")
                    fields_output = fields_output + key + ":" + str(value) + ", "
                fields_output = str(i) + " - " + fields_output[:len(fields_output)-2] + "}"
                print(fields_output)
            self.log_equal_action("Get JSON from API", "n/a", "n/a", description)
        else:
            print(f"Failed to retrieve API data. Status code: {response.status_code}")
            self.log_equal_action("Get JSON from API", "n/a", "Failed", description)


    def log_equal_action(self, action: str, expected: str, actual: str, description: str) -> None:
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        # print_line(f"In log_equal_action workbook = {workbook.path}, sheet = {sheet.title}")
        status = "Fail"
        now = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        if expected == actual:
            status = "Pass"
        row = self.find_first_empty_row()
        if row == 1:
            self.print_headers(sheet)
            row = row + 1
        sheet.cell(row, 1).value = action
        sheet.cell(row, 2).value = expected
        sheet.cell(row, 3).value = actual
        sheet.cell(row, 4).value = status
        sheet.cell(row, 5).value = str(now)
        sheet.cell(row, 6).value = description

        if status == "Pass":
            sheet.cell(row, 4).style = 'Good'
        else:
            sheet.cell(row, 4).style = 'Bad'

        # need to account for self.driver being unavailable because quit() was already initiated
        # if self.driver.session_id is None:
        if action.lower() == "close driver":
            sheet.cell(row, 7).value = "n/a"
            self.set_close_driver_theme(sheet, row)
            print("\n")
        else:
            try:
                if self.driver.current_url != "data:,":
                    sheet.cell(row, 7).value = self.driver.current_url
                elif self.check_response_url is not None and len(self.check_response_url) > 0:
                    sheet.cell(row, 7).value = self.check_response_url
                else:
                    sheet.cell(row, 7).value = "Non-Site Related Command"
            except Exception as e:
                print(f"{variables.terminal_color_red}Error getting current_url: {e}{variables.terminal_color_reset}")
                sheet.cell(row, 7).value = "Error getting current_url"
        workbook.save(self.log_file_name)
        print("-" * variables.dash_length)
        # print(f"Action: {action}{self.get_print_spacing(True, action)}\t|\tDescription:{description}{self.get_print_spacing(True, description)}\t|\tStatus:{status}")
        if status == "Pass":
            print(f"{variables.terminal_color_green}Action: {action}{self.get_print_spacing(True, action)}\t|\tDescription:{description}{self.get_print_spacing(True,description)}\t|\tStatus:{status}{variables.terminal_color_reset}")
        else:
            print(f"{variables.terminal_color_red}Action: {action}{self.get_print_spacing(True, action)}\t|\tDescription:{description}{self.get_print_spacing(True, description)}\t|\tStatus:{status}{variables.terminal_color_reset}")
        print("-" * variables.dash_length)

    def log_contains_action(self, action, expected, actual, description):
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        now = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        status = self.get_status_contains(expected, actual)
        row = self.find_first_empty_row()
        if row == 1:
            self.print_headers(sheet)
            row = row + 1
        sheet.cell(row, 1).value = action
        sheet.cell(row, 2).value = expected
        sheet.cell(row, 3).value = actual
        sheet.cell(row, 4).value = status
        sheet.cell(row, 5).value = str(now)
        sheet.cell(row, 6).value = description

        if status == "Pass":
            sheet.cell(row, 4).style = 'Good'

        else:
            sheet.cell(row, 4).style = 'Bad'
        # need to account for self.driver being unavailable because quit() was already initiated
        # if self.driver.session_id is None:
        if action.lower() == "close driver":
            print("...")
            sheet.cell(row, 7).value = "n/a"
            self.set_close_driver_theme(sheet, row)
            print(" \n ")
        else:
            sheet.cell(row, 7).value = self.driver.current_url
        workbook.save(self.log_file_name)
        print(f"Action: {action}\t|\tDescription:{description}\t|\tStatus:{status}")

    # This method finds the first empty row in the Excel Spreadsheet
    def find_first_empty_row(self) -> int:
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        row = 1
        while sheet.cell(row, 1).value is not None and len(sheet.cell(row, 1).value) > 0:
            row = row + 1

        return row

    @staticmethod
    def get_print_spacing(is_action, in_text):
        min_action_len = 25
        min_desc_len = 50
        action_spaces = (min_action_len - len(in_text)) * "."
        desc_spaces = (min_desc_len - len(in_text)) * "."
        if is_action:
            return action_spaces
        else:
            return desc_spaces

    # This method gets the xPath and Css Selector values for all elements matching the
    # accessor_type and accessor combination and supplies the tag_name to help with
    # referencing the item.
    def get_all_element_xpath_css_values(self, accessor_type:str, accessor:str, file_name:str, description:str) -> None:
        if ":" not in file_name and "selector_files/" not in file_name:
            file_name = "selector_files/" + file_name
        file = open(file_name, "w")
        file.write(f"The xPath and Css Selectors are for URL: {self.driver.current_url}\r\n")
        file.write(f"Selector Type: {accessor_type} - Selector: {accessor}\r\n")
        file.write("-" * variables.dash_length)
        file.write(f"\r\n")
        status = False
        elements = self.get_elements(accessor_type, accessor)

        for element in elements:
            # print(f"\nTag_Name = {element.tag_name}")
            file.write(f"Tag_Name = {element.tag_name}\r\n")
            if element.get_attribute("id") is not None and len(element.get_attribute("id")) > 0:
                xpath = f"XPath = //{element.tag_name}" + "[@id='" + element.get_attribute("id") + "']"
                # print(f"XPath = {xpath}")
                file.write(f"{xpath}\r\n")
                # print(f'Css_Selector = {element.tag_name}#{element.get_attribute("id")}')
                file.write(f'Css_Selector = {element.tag_name}#{element.get_attribute("id")}\r\n')
                status = True
            elif element.get_attribute("class") is not None and len(element.get_attribute("class")) > 0:
                xpath = f"XPath = //{element.tag_name}" + "[@class='" + element.get_attribute("class") + "']"
                # print(f"XPath = {xpath}")
                file.write(f"{xpath}\r\n")
                # print(f'Css_Selector = {element.tag_name}.{element.get_attribute("class").replace(" ",".")}')
                file.write(f'Css_Selector = {element.tag_name}.{element.get_attribute("class").replace(" ",".")}\r\n')
                status = True
            else:
                parent = element.find_element(By.XPATH, "..")
                xpath_selector = element.tag_name
                css_selector = element.tag_name
                while parent is not None:
                    xpath_selector = parent.tag_name + "/" + xpath_selector
                    # css_selector = parent.tag_name + parent.get_attribute("class").replace(' ','.') + " > " + css_selector
                    css_selector = parent.tag_name + " > " + css_selector
                    if parent.tag_name == "html":
                        break
                    parent = parent.find_element(By.XPATH, "..")
                    # print(f"In process parent.tag_name = {parent.tag_name} and xpath_selector = {xpath_selector}")
                xpath_selector = "//" + xpath_selector
                # print(f"XPath = {xpath_selector}")
                file.write(f"XPath = {xpath_selector}\r\n")
                # print(f"Css_Selector = {css_selector}")
                file.write(f"Css_Selector = {css_selector}\r\n")
                status = True
        if status:
            print(f"\r\nxPath and Css Selectors written to file: {file_name}")
        else:
            print(f"\r\nSomething went wrong when attempting to get xPath and Css Selectors for file: {file_name}")
        # self.log_equal_action("Get All Element xPaths & Css Selectors", str(True), str(status), description)
        self.log_equal_action("Get xPaths & Css Selectors", str(True), str(status), description)

    def get_element_xpath(self, element):
        if element.get_attribute("id") is not None and len(element.get_attribute("id")) > 0:
            xpath = f"//{element.tag_name}" + "[@id='" + element.get_attribute("id") + "']"
        else:
            parent = element.find_element(By.XPATH, "..")
            xpath = element.tag_name
            while parent is not None:
                xpath = parent.tag_name + "/" + xpath
                if parent.tag_name == "html":
                    break
                parent = parent.find_element(By.XPATH, "..")
            xpath = "//" + xpath
        return xpath

    # This method prints all table information based on the accessor_type and accessor, but
    # can also change the display orientation based on the text/URL value.
    # The default orientation is vertical.
    # def get_table_information(self, accessor_type, accessor, file_name, description):
    def get_table_information(self, accessor_type:str, accessor:str, display_orientation:str, description:str) -> None:
        if len(display_orientation) <= 0:
            display_orientation = "vertical"
        table = self.get_element(accessor_type, accessor)
        table_rows = table.find_elements(By.TAG_NAME, "tr")
        table_headers = ""
        row_data = ""
        for tr in table_rows:
            if table_headers == "":
                table_headers = tr.find_elements(By.TAG_NAME, "th")
                for th in table_headers:
                    if display_orientation == "vertical":
                        print(th.text)
                    else:
                        row_data = row_data + th.text + "\t| "
                if len(row_data) > 0:
                    print(row_data)
            table_cells = tr.find_elements(By.TAG_NAME, "td")
            row_data = ""
            for td in table_cells:
                if display_orientation == "vertical":
                    print(td.text)
                else:
                    row_data = row_data + td.text + "\t| "
            if len(row_data) > 0:
                print(row_data)
        self.log_equal_action("Get Table Data", "n/a", "n/a", description)

    # This method prints all table information based on the accessor_type and accessor, but
    # can also change the display orientation based on the text/URL value.
    # The default orientation is vertical.
    # This method also allows the test to specify the minimum space each column should
    # take up, using the Expected value, so that the output aligns like a table should.
    def get_table_information_alt(self, accessor_type, accessor, display_orientation, max_space_str, description):
        if len(display_orientation) <= 0:
            display_orientation = "vertical"
        table = self.get_element(accessor_type, accessor)
        table_rows = table.find_elements(By.TAG_NAME, "tr")
        table_headers = ""
        row_data = ""
        max_space = int(max_space_str)
        for tr in table_rows:
            if table_headers == "":
                table_headers = tr.find_elements(By.TAG_NAME, "th")
                for th in table_headers:
                    if display_orientation == "vertical":
                        print(th.text)
                    else:
                        row_data = row_data + th.text + ("." * (max_space - len(th.text)))  + "\t| "
                if len(row_data) > 0:
                    print(row_data)
            table_cells = tr.find_elements(By.TAG_NAME, "td")
            row_data = ""
            for td in table_cells:
                if display_orientation == "vertical":
                    print(td.text)
                else:
                    row_data = row_data + td.text + ("." * (max_space - len(td.text))) + "\t| "

            if len(row_data) > 0:
                print(row_data)
        self.log_equal_action("Get Table Data", "n/a", "n/a", description)

    # This method saves each page's HAR file to the global har contents variable
    # so that tag information can be extracted at a later time.
    def get_har_file_for_tag_information(self) -> None:
        if self.driver.current_url == "data:,":
            return
        java_script_command = "const resources = performance.getEntriesByType('resource');\n"
        java_script_command = java_script_command + "var returnValue = '';\nresources.forEach((entry) => {\n"
        java_script_command = java_script_command + "   returnValue += `${entry.name}'s startTime: ${entry.startTime}\r\n`\n"
        java_script_command = java_script_command + "});\nreturn returnValue;"
        status = False
        description = "Unsolicited HAR retrieval for tagging information"
        har_content = str(self.driver.execute_script(java_script_command))
        if har_content is not None and len(har_content) > 0:
            # save all har file content to the class level variable
            self.har_file_contents = self.har_file_contents + har_content
            status = True
            # If this doesn't have a save complete har file among the commands,
            # get the GA4 tags in the har content

            if not self.check_save_complete_har_file():
                self.get_ga4_analytics_tags(har_content)
                # self.get_ga4_analytics_tags(har_content, "")
        self.log_equal_action("Get HAR for tag data.", str(True), str(status), description)

    def save_complete_har_file(self, file_name:str = None, description:str = None) -> None:
        java_script_command = "const resources = performance.getEntriesByType('resource');\n"
        java_script_command = java_script_command + "var returnValue = '';\nresources.forEach((entry) => {\n"
        java_script_command = java_script_command + "   returnValue += `${entry.name}'s startTime: ${entry.startTime}\r\n`\n"
        java_script_command = java_script_command + "});\nreturn returnValue;"
        status = False
        file_name = self.update_file_name(file_name,"har")
        # If the har_file_name is blank, it is the first time through so delete the old file
        # for subsequent saves, just append to the existing file
        if self.har_file_name == "" or self.har_file_name is None:
            self.delete_file(file_name, "Complete Har File")
            self.har_file_name = file_name

        har_content = str(self.driver.execute_script(java_script_command))
        self.har_file_contents = self.har_file_contents + har_content
        if self.har_file_contents is not None and len(self.har_file_contents) > 0:
            # if no file_name is passed in, create one
            # if file_name is None or len(file_name) <= 0:
            #     now = datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
            #     file_name = "har_file" + now + ".txt"
            # If file_name not being set to specific file location and doesn't include the data folder
            # add the data folder, otherwise leave the file_name alone.
            # if ":" not in file_name and "har_files/" not in file_name:
            #     file_name = "har_files/" + file_name
            file = open(file_name, "a")
            file.write(self.har_file_contents)
            print(f"har_content written to ({file_name}).\r\n")
            status = True
            # self.get_ga4_analytics_tags(self.har_file_contents, file_name)
            self.get_ga4_analytics_tags(self.har_file_contents)
        self.log_equal_action("Save HAR file", str(True), str(status), description)


    # Thinking of removing this method and just keeping the save_complete_har_file
    # in which I can add the page url at the top of the file to delineate between different URLs
    def save_har_file(self, file_name:str = None, description:str = None) -> None:
        java_script_command = "const resources = performance.getEntriesByType('resource');\n"
        java_script_command = java_script_command + "var returnValue = '';\nresources.forEach((entry) => {\n"
        java_script_command = java_script_command + "   returnValue += `${entry.name}'s startTime: ${entry.startTime}\r\n`\n"
        java_script_command = java_script_command + "});\nreturn returnValue;"
        status = False
        file_name = self.update_file_name(file_name, "har")
        # If the har_file_name is blank, it is the first time through so delete the old file
        # for subsequent saves, just append to the existing file
        if self.har_file_name == "" or self.har_file_name is None:
            self.delete_file(file_name, "Individual Har File")
            self.har_file_name = file_name

        har_content = str(self.driver.execute_script(java_script_command))

        if har_content is not None and len(har_content) > 0:
            # if no file_name is passed in, create one
            # if file_name is None or len(file_name) <= 0:
            #     now = datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
            #     file_name = "har_file" + now + ".txt"
            # If file_name not being set to specific file location and doesn't include the data folder
            # add the data folder, otherwise leave the file_name alone.
            # if ":" not in file_name and "har_files/" not in file_name:
                # file_name = "har_files/" + file_name
            file = open(file_name, "a")
            file.write(har_content)
            # print(f"har_content written to ({file_name}).\r\n{har_content}")
            print(f"har_content written to ({file_name}).\r\n")
            status = True
            # self.get_ga4_analytics_tags(har_content, file_name)
            self.get_ga4_analytics_tags(har_content)
        self.log_equal_action("Save HAR file", str(True), str(status), description)

    # def get_ga4_analytics_tags(self, har_content:str, file_name:str) -> None:
    def get_ga4_analytics_tags(self, har_content:str) -> None:
        # May eventually allow saving tags to a file but currently, this just prints them out
        status = False
        description = "Check for GA4 Analytics Tags"
        # file_name = file_name.replace(".", "_GA4_Tags.")
        if variables.ga4_analytics_identifier in har_content and variables.google_analytics_identifier in har_content:
            # if self.has_save_complete:
            print("GA4 Analytics Tags Found!")
            ar_har_content = har_content.split("\n")
            for har_entry in ar_har_content:
                if variables.ga4_analytics_identifier in har_entry:
                    ar_har_values = urlparse(har_entry)
                    params = parse_qs(ar_har_values.query)
                    event_name = params.get("en", [None])[0]
                    document_location = params.get("dl", [None])[0]
                    hit_type = params.get("t", [None])[0]
                    tracking_id = params.get("tid", [None])[0]
                    # if tracking_id is not None and event_name is not None:
                    if tracking_id is not None:
                        print(f"Tracking Id:{tracking_id} - Event Name: {event_name} - Hit Type:{hit_type} - Document Location:{document_location}")
                    status = True
            self.log_equal_action("GA4 Analytics Tags Found", str(True), str(status), description)
        else:
            print("No GA4 Analytics Tags Found")
        # self.log_equal_action("GA4 Analytics Tags Found", str(True), str(status), description)




    # This is just an idea at this point but the gist is this:
    # the selector_type has a tag type, i.e.: GA4, TikTok, facebook, pinterest
    # the selector field has a comma-delimited string containing values to
    # check for so this loops through the har_content and searches for the
    # information.  The expected field contains a comma-delimited list of
    # tag parameters to match against and the text_url field is the har_content or har file.
    # def check_page_tagging(self, selector_type, selector, text_url, expected, description):
    def check_page_tagging(self, selector_type, text_url, expected, description):
        if text_url is not None and len(text_url) > 1 and os.path.exists(text_url):
            har_contents = ""
            file = open(text_url,"r")
            for line in file:
                har_contents = har_contents + line + "\n"
            ar_har = har_contents.split("\n")
        elif text_url is not None and len(text_url) > 1 and not os.path.exists(text_url):
            print(f"The supplied path does not exist: {text_url}")
            return
        else:
            if len(self.har_file_contents) <= 0:
                self.get_har_file_for_tag_information()
            ar_har = self.har_file_contents.split("\n")
            # print("Har contents: " + self.har_file_contents)
        ar_expected = expected.split(",")
        items = len(ar_expected)
        found = 0
        for har_entry in ar_har:
            # print(f"Checking har Entry: {har_entry} for selector_type: {selector_type}")
            if selector_type in har_entry:
                for exp in ar_expected:
                    # remove any spaces in the search expression which should be key=value
                    exp = exp.replace(' ', '')
                    if exp in har_entry:
                        found = found + 1
                        print(f"Found selector_type: {selector_type} and value: {exp}")
                        if found == items:
                            print("All tag parameters found!")
                            break
            if found == items:
                break
        status = True if items == found else False
        if status:
            status_message = f"Matching {selector_type} tag found!"
        else:
            status_message = f"Matching {selector_type} tag NOT found!"
        print(status_message)
        self.log_equal_action(f"{selector_type} tag found", str(True), str(status), description)

    def check_image_tags_for_alt_text(self, selector_type, selector, description):
        elements = self.get_elements(selector_type, selector)
        total_count = len(elements)
        missing_alt_count = 0
        for element in elements:
            alt_text = element.get_attribute("alt")
            if alt_text is None or alt_text == nullcontext or len(alt_text) <= 0:
                print(f"Element missing alt text: {element.get_attribute('outerHTML')}")
                missing_alt_count = missing_alt_count + 1
        # status = True if missing_alt_count == 0 else False
        self.log_equal_action(f"Image Alt Tag Check", str(total_count), str(missing_alt_count), description)

    # def wcag_ada_check_controller_old(self, description):
    #     # overall selector_type
    #     selector_type = "css_selector"
    #     # image element
    #     selector = variables.wcag_ada_img[0]
    #     expected = variables.wcag_ada_img[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # wcag_ada_button element
    #     selector = variables.wcag_ada_button[0]
    #     expected = variables.wcag_ada_button[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # nav element
    #     selector = variables.wcag_ada_nav[0]
    #     expected = variables.wcag_ada_nav[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # section element
    #     selector = variables.wcag_ada_section[0]
    #     expected = variables.wcag_ada_section[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # form element
    #     selector = variables.wcag_ada_form[0]
    #     expected = variables.wcag_ada_form[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # table
    #     selector = variables.wcag_ada_table[0]
    #     expected = variables.wcag_ada_table[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     #th element
    #     selector = variables.wcag_ada_th[0]
    #     expected = variables.wcag_ada_th[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # iframe element
    #     selector = variables.wcag_ada_iframe[0]
    #     expected = variables.wcag_ada_iframe[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # svg element
    #     selector = variables.wcag_ada_svg[0]
    #     expected = variables.wcag_ada_svg[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # video element
    #     selector = variables.wcag_ada_video[0]
    #     expected = variables.wcag_ada_video[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # source element
    #     selector = variables.wcag_ada_video_source[0]
    #     expected = variables.wcag_ada_video_source[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)
    #     # track element
    #     selector = variables.wcag_ada_video_track[0]
    #     expected = variables.wcag_ada_video_track[1:]
    #     self.wcag_ada_checks(selector_type, selector, expected, description, True)

    def wcag_ada_check_controller(self, description):
        selector_type = "css_selector"
        for item in variables.wcag_ada_array:
            selector = item[0]
            expected = item[1:]
            self.wcag_ada_checks(selector_type, selector, expected, description, True)

    def wcag_ada_checks(self, selector_type, selector, expected, description, is_array = False):
        log_action = selector if selector is not None else "Complete"
        message_printed = False
        if selector_type is not None:
            print(selector_type, selector, is_array)
            elements = self.get_elements(selector_type, selector)
            if not is_array:
                items = expected.split(",")
            else:
                items = expected
            total_count = len(elements) * len(items)
            print(f"Total elements found for {selector}: {total_count}")
            missing_param_count = total_count
            for element in elements:
                for item in items:
                    if "<" in item:
                        child_selector = item.replace("<","").replace(">","")
                        print(f"child_selector = {child_selector}")
                        try:
                            child = element.find_element(By.TAG_NAME, child_selector)
                            print(f"Child element is: {child}")
                            if child is None or child == nullcontext or len(child) <= 0:
                                print(f"Element missing child ({child}): {element.get_attribute('outerHTML')}")
                                missing_param_count = missing_param_count - 1
                        except:
                            print(f"Child element {child_selector} does not exist.")
                    else:
                        if element.tag_name.lower() != "a":
                            param = element.get_attribute(item)
                            print(f"Checking for param = {item}")
                            if param is None or param == nullcontext or len(param) <= 0:
                                print(f"Element missing {item}: {element.get_attribute('outerHTML')}")
                                missing_param_count = missing_param_count - 1
                        else:
                            if not message_printed:
                                print(f"Checking for invalid text: {item}")
                                message_printed = True
                            check_text = self.get_element_text_silently(element)
                            if check_text.lower() == item.lower():
                                print(f"Element has invalid text: {item}: {element.get_attribute('outerHTML')}")
                                missing_param_count = missing_param_count - 1

            self.log_equal_action(f"{log_action} WCAG/ADA Check", str(total_count), str(missing_param_count), description)

    # Placed all save file type updating for directory information into a single method
    # where the file_type parameter determines where the file will be saved
    @staticmethod
    def update_file_name(file_name, file_type):
        if file_type.lower() == "har" and ":" not in file_name and "har_files/" not in file_name:
            if file_name is None or len(file_name) <= 0:
                now = datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
                file_name = "har_file" + now + ".txt"
            return "har_files/" + file_name
        elif file_type.lower() == "selector" and ":" not in file_name and "selector_files/" not in file_name:
            return "selector_files/" + file_name
        elif file_type.lower() == "screenshot" and ":" not in file_name and "screenshots/" not in file_name:
            return "screenshots/" + file_name
        return file_name

    # This method deletes the file based on the file_name parameter passed in
    @staticmethod
    def delete_file(file_name, description = None) -> bool:
        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"Prior {description} ({file_name}) - File deleted.")
            return True
        else:
            print(f"Prior {description} ({file_name}) - File does not exist.")
            return False



    @staticmethod
    def print_headers(sheet):
        row = 1
        start_style = "Neutral"
        sheet.cell(row, 1).value = "Action"
        sheet.cell(row, 2).value = "Expected"
        sheet.cell(row, 3).value = "Actual"
        sheet.cell(row, 4).value = "Status"
        sheet.cell(row, 5).value = "Date and Time Executed"
        sheet.cell(row, 6).value = "Description"
        sheet.cell(row, 7).value = "Page URL"
        sheet.cell(row, 1).style = start_style
        sheet.cell(row, 2).style = start_style
        sheet.cell(row, 3).style = start_style
        sheet.cell(row, 4).style = start_style
        sheet.cell(row, 5).style = start_style
        sheet.cell(row, 6).style = start_style
        sheet.cell(row, 7).style = start_style

    @staticmethod
    def set_close_driver_theme(sheet, row):
        end_style = "Neutral"
        sheet.cell(row, 1).style = end_style
        sheet.cell(row, 2).style = end_style
        sheet.cell(row, 3).style = end_style
        sheet.cell(row, 4).style = end_style
        sheet.cell(row, 5).style = end_style
        sheet.cell(row, 6).style = end_style
        sheet.cell(row, 7).style = end_style

    @staticmethod
    def get_status_contains(contained_string, full_string):
        print(f"contained_string = {contained_string}")
        print(f"full_string = {full_string}")
        if contained_string in full_string:
            return "Pass"
        else:
            return "Fail"

    def read_excel_command_file(self, test_file = None, verbose = True):
        if verbose:
            print(f"Reading Commands File:{self.command_file}")
        self.command_file = test_file if test_file is not None else self.command_file
        workbook = openpyxl.load_workbook(self.command_file)
        sheet = workbook[self.command_sheet]
        self.data = []
        for row in range(1, sheet.max_row + 1):
            command = sheet.cell(row, 1).value
            selector_type = sheet.cell(row, 2).value
            selector = sheet.cell(row, 3).value
            text_url = sheet.cell(row, 4).value
            expected = sheet.cell(row, 5).value
            actual = sheet.cell(row, 6).value
            description = sheet.cell(row, 7).value
            self.data.append((command, selector_type, selector, text_url, expected, actual, description))
        return self.data

    # Checks if there is a Save complete har file command, if found returns True, else False
    def check_save_complete_har_file(self):
        self.has_save_complete = False
        command_list = self.read_excel_command_file(self.command_file, False)
        for command, selector_type, selector, text_url, expected, actual, description in command_list:
            if command.lower() == command_library.save_complete_har_file:
                self.has_save_complete = True
                break

        return self.has_save_complete

    # This is the way you get the second-highest number from an Array of numbers.
    # First, get a unique set of numbers, the set function does this.
    # Next, sort the unique numbers from lowest to highest,
    # Then place them in reverse order so they are highest to lowest.
    # Check to ensure that more than 1 number is in the list otherwise, only the highest is remaining.
    # Finally, get the second item from the array or list and that is the second-highest number.
    @staticmethod
    def get_second_highest_number_from_array():
        numbers = [10, 20, 40, 20, 50, 40, 50]

        # Remove duplicates and sort in descending order
        unique_numbers = sorted(set(numbers), reverse=True)
        print("unique_numbers = ", unique_numbers)

        # Check that there are at least 2 unique numbers
        if len(unique_numbers) >= 2:
            second_highest = unique_numbers[1]
            print("Second highest number is:", second_highest)
        else:
            print("Not enough unique numbers to determine second highest.")

    def connect_to_database(self, text_url, description):
        connection_objects = text_url.split(",")
        server_objects = connection_objects[0].split("=")
        server = server_objects[1]
        database_objects = connection_objects[1].split("=")
        database = database_objects[1]
        print(f"server = {server}")
        print(f"database = {database}")
        connection_string = variables.connection_string_template.replace("*server*", server).replace("*database*", database)
        print(f"connection_string = {connection_string}")
        self.engine = create_engine(connection_string)
        status = False
        if self.engine is not None:
            status = True
        self.log_equal_action("Database Connection", str(True), str(status), description)
        return connection_string

    def get_data_using_sql_alchemy(self, sql_query, connection_string, description, show_data = False):
        if self.engine is None:
            print(f"Database engine not set in the constructor!\nAttempting to set it using connection_string: {connection_string}")
            self.engine = create_engine(connection_string)
            if self.engine is None:
                print("Database engine is still not set! Exiting get_data() function!")
                return
        if sql_query is not None:
            with self.engine.connect() as conn:
                with suppress_stderr():
                    data = pandas.read_sql(sql_query, conn)
                    print(data)
                    if show_data:
                        for index, row in data.iterrows():
                            print(f"Row {index}:")
                            for col in data.columns:
                                print(f"  {col}: {row[col]}")
                        print("-" * 100)
                status = True if data is not None else False
                self.log_equal_action("Query Database", str(True), str(status), description)
                # return data



    def __del__(self):
        pass
        # self.engine

    def perform_right_click(self, selector_type, selector, description):
        # action_chains = ActionChains(self.driver)
        # element = self.get_element(selector_type, selector)
        # action_chains.context_click(element).perform()
        # self.log_equal_action("Right Click", "n/a", "n/a", description)
        pass

    def open_link_in_new_tab(self, selector_type, selector, description):
        print(f"in open_link_in_new_tab selector_type: {selector_type} selector:{selector}")
        context_menu_option = self.get_element(selector_type, selector)
        print(f"in open_link_in_new_tab selector_type: sending keys")
        context_menu_option.send_keys(keys.Keys.CONTROL + keys.Keys.RETURN)
        print(f"in open_link_in_new_tab selector_type: keys sent")
        self.log_equal_action("Context Click", "n/a", "n/a", description)

    def close_tab(self, index, description):
        win_handle = self.driver.window_handles
        print(f"win_handle = {win_handle}")
        current_handle = self.driver.current_window_handle
        current_index = 0
        for handle in win_handle:
            if handle == current_handle:
                break
            else:
                current_index += 1
        print(f"index = {index} and current_index = {current_index}")
        # status = False
        try:
            self.driver.switch_to.window(win_handle[int(index)])
            print(f"switched to tab to close")
            self.driver.close()
            print(f"closed tab")
            status = True
        except WebDriverException as w:
            status = False
            print(f"Closed Tab WebDriverException {w}")
        except Exception as e:
            status = False
            print(f"Closed Tab Exception {e}")
        self.driver.switch_to.window(win_handle[current_index])
        self.log_equal_action("Close Tab", str(True), str(status), description)

    # This method begins with the url provided in the text_url parameter,
    # gets a list of all a tag elements, gets the domain value from the
    # text_url parameter, loops through those links calling a method to
    # retrieve site URLs from the element hrefs, and add to the list of
    # links and when done navigating all pages, saves the list to a
    # text file provided by the expected field, or it alters the URL to
    # make a valid file name.
    # NOTE: if a selector is provided to narrow down
    # the type of links returned, it is used, else all a tags are retrieved.
    # def spider_site(self, selector_type, selector, text_url, expected, description):
    def spider_site(self, selector, text_url, expected, description):
        links = []
        status = "No Links"
        nav_description = "Spidering Site"
        domain = urlparse(text_url).netloc
        file_name = expected if expected is not None else text_url
        self.navigate(text_url, nav_description)
        links.append(text_url)
        time.sleep(3)
        accessor_type = "tag_name"
        accessor = "a" if selector is None else selector
        try:
            elements = self.get_elements(accessor_type, accessor)
            links = self.get_all_hrefs(elements, links, domain)
            for link in links:
                print(f"In for loop going to {link}")
                self.navigate_without_checking(link,nav_description)
                time.sleep(3)
                elements = self.get_elements(accessor_type, accessor)
                links = self.get_all_hrefs(elements, links, domain)
            if len(links) > 0:
                status = "Links"
                file_content = ','.join(links).replace(',','\n')
                self.wps.save_to_file(file_name,file_content)
            else:
                status = "No Links"
        except Exception as e:
            file_content = ','.join(links).replace(',', '\n')
            self.wps.save_to_file(file_name, file_content)
            print(f"Spider Exception: {e}")
        self.log_equal_action("Spider Site", "Links",status, description)

    # This method gets all link href urls from all a tag elements passed in,
    # if they match the domain, do not contain javascript and do not contain mailto
    # and returns the list of links to the calling method
    def get_all_hrefs(self, elements, links, domain):
        link_exists = False
        for element in elements:
            href = element.get_attribute("href") if element.get_attribute("href") is not None else None
            print(f"href = {href}")
            if href is not None and href.find(domain) >= 0 > href.lower().find(
                    'javascript') and href.lower().find('mailto:') < 0:
                for link in links:
                    if href == link:
                        link_exists = True
                if not link_exists:
                    links.append(element.get_attribute("href"))
            link_exists = False
        return links

    def check_response_code(self, page_url, expected, description):
        if expected is None:
            expected = "200"
        try:
            response = requests.get(page_url)
            time.sleep(2)
            print(f" Status Code: Expected: {expected} Actual: {response.status_code}")
            actual = str(response.status_code)
        except requests.exceptions.RequestException as r:
            actual = "Exception"
            print(f" Status Code: Expected: {expected} Actual: {variables.terminal_color_red} {actual}: {r} {variables.terminal_color_reset}")
        self.check_response_url = page_url
        self.log_equal_action("Check Response Code", expected, actual, description)
        self.check_response_url = ""

    def check_response_codes(self, text_url, expected, description):
        if ".txt" in text_url:
            data = self.read_text_file(text_url)
        elif ".xlsx" in text_url:
            data = self.read_excel_data_file(text_url)
        else:
            print("No method exists for reading this type of file!  \nMake a .txt or .xlsx file to use this command.")
            return
        for item in data:
            print(f"Check response code for:{item}")
            if expected is None:
                expected = "200"
            self.check_response_code(item, expected, description)

    # This method reads URLs from a Text file
    # and returns a list to the calling method
    def read_text_file(self, text_url):
        data = []
        if not "\\" in text_url and not "/" in text_url:
            text_url = "./data/" + text_url

        with open(text_url, 'r') as file:
            for line in file:
                if line is not None and len(line) > 0:
                    data.append(line)
                    # print(f"line = {line}")
        return data

    # This method reads URLs from an Excel file
    # and returns a list to the calling method
    def read_excel_data_file(self, text_url):
        if not "\\" in text_url and not "/" in text_url:
            text_url = "./data/" + text_url
        print(f"in read_excel_data_file - text_url = {text_url}")
        workbook = openpyxl.load_workbook(text_url)
        sheet = workbook.worksheets[0]
        print(f"Attempting to read worksheet: {sheet.title}")
        data = []
        for row in range(1, sheet.max_row + 1):
            line = sheet.cell(row, 1).value
            if line is not None and len(line) > 0:
                data.append(line)
        return data