import os
import random
import time

import openpyxl
from selenium.common import ElementNotInteractableException, StaleElementReferenceException

from common.function_library import Functions
from common import variables
from common import command_library
from openpyxl import Workbook
from urllib.parse import urlparse, parse_qs

class CreateTestFile:
    def __init__(self, funct):
        self.funct = Functions(funct.driver)
        self.funct = funct
        self.file_name = ""


    def create_test_file_controller(self, selector_type, selector, text_url, expected, description):
        if selector_type is not None and selector is not None:
            accessor_type = selector_type
            accessor = selector
        else:
            accessor_type = "css_selector"
            accessor = "*"
        self.file_name = text_url
        if not "\\" in self.file_name and not "/" in self.file_name:
            self.file_name = "./data/" + self.file_name
        self.check_file_name_unique()
        self.create_navigation_command(self.funct.driver.current_url)
        self.check_for_one_trust_consent()
        self.create_screenshot_command()
        self.create_wcag_command(self.funct.driver.current_url)
        element = None
        try:
            elements = self.funct.get_elements(accessor_type, accessor)
            for element in elements:
                try:
                    if element.tag_name == "a":
                        self.create_command_check_response_code(element)
                        print(f"Tag Type: {element.tag_name}\n{element.get_attribute('href')}")
                        print(f"Command:Check Response Status Code, text_url: {element.get_attribute('href')}, expected: 200")
                    elif (element.tag_name == "input" and element.get_attribute("type") == "text") or element.tag_name == "textarea":
                        try:
                            element.click()
                            self.create_command_send_keys(element)
                        except ElementNotInteractableException as i:
                            print(f"Element {element.tag_name} - {element} is not interactable, skipping.")
                    elif element.tag_name == "input" and (element.get_attribute("type") == "checkbox" or element.get_attribute("type") == "radio"):
                        self.create_click_command(element)
                        print(f"Tag Type: {element.tag_name}\nType:{element.get_attribute("type")}")
                    elif element.tag_name == "select":
                        self.create_select_command(element)
                    elif element.tag_name == "img":
                        self.create_command_check_response_code(element)
                        print(f"Tag Type: {element.tag_name}\n{element.get_attribute('src')}")
                        print(f"Command:Check Response Status Code, text_url: {element.get_attribute('src')}, expected: 200")
                    elif element.tag_name == "h1":
                        self.create_command_check_header_value(element)
                        print(f"Tag Type: {element.tag_name}\nValue:{self.funct.get_element_text_silently(element)}")
                except StaleElementReferenceException as se:
                    print(f"Stale Element : \n{se}")
            print("Sleeping for 5 seconds to allow tagging to fire.")
            self.funct.get_har_file_for_tag_information()
            time.sleep(5)
            if self.funct.har_file_contents is not None and len(self.funct.har_file_contents) > 0:
                self.check_for_analytics_tagging()
                self.check_for_media_tagging()
        except Exception as e:
            if element is not None:
                tag_name = element.tag_name
            else:
                tag_name = "Tag Name irretrievable"
            print(f"Error attempting to create command for tag_name:{tag_name}\n{e}")
                # print(f"Command:Check Response Status Code, Selector_Type: {element.tag_name}, Selector: {element.get_attribute('href')}")
                # print(f"Tag Type: {element.tag_name}\n{element.get_attribute('outerHTML')}")

    #This method creates the Navigation command
    def create_navigation_command(self, current_url):
        command = command_library.navigate
        selector_type = ""
        selector = ""
        text_url = current_url
        expected = current_url
        actual = "current_url"
        description = "Navigate to URL"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    def create_command_check_header_value(self, element):
        if element is not None and element.tag_name == "h1":
            command = "Get Text"
            selector_type = "tag_name"
            selector = "h1"
            text_url = ""
            expected = self.funct.get_element_text_silently(element)
            actual = ""
            description = "Get H1 Text"
            self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    def create_command_check_response_code(self, element):
        command = command_library.check_response_status_code
        selector_type = "" #"xpath"
        selector = "" #self.funct.get_element_xpath(element)
        if element.tag_name == "a":
            text_url = element.get_attribute('href')
        else:
            text_url = element.get_attribute('src')
        expected = "200"
        actual = ""
        description = f"Check {element.tag_name} Tag Response Status code"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    # This method creates the send keys command for text boxes
    def create_command_send_keys(self, element):
        command = command_library.send_keys
        selector_type = "xpath"
        selector = self.funct.get_element_xpath(element)
        text_url = "Phrase"
        expected = ""
        actual = ""
        description = "Send keystrokes to the element"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    # This method creates the WCAG Command to test all WCAG element compliances
    def create_wcag_command(self, current_url):
        command = command_library.perform_wcag_ada_checks
        selector_type = ""
        selector = ""
        text_url = ""
        expected = ""
        actual = ""
        description = "Check all WCAG & ADA HTML elements "
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    #This method writes the command to the Test file
    def write_command_to_file(self, command, selector_type, selector, text_url, expected, actual, description):
        if not os.path.exists(self.file_name):
            self.create_new_workbook()
        workbook = openpyxl.load_workbook(self.file_name)
        ws = workbook["CommandScripts"]
        row = 2
        while ws.cell(row, 1).value is not None and len(ws.cell(row, 1).value) > 0:
            row = row + 1

        ws.cell(row, 1).value = command
        ws.cell(row, 2).value = selector_type
        ws.cell(row, 3).value = selector
        ws.cell(row, 4).value = text_url
        ws.cell(row, 5).value = expected
        ws.cell(row, 5).number_format = "@"
        ws.cell(row, 6).value = actual
        ws.cell(row, 7).value = description
        workbook.save(self.file_name)



    # This method creates a new workbook, names the sheet
    # CommandScripts and adds the header names
    def create_new_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "CommandScripts"
        # ws.append(["print('Hello')", "Prints Hello to screen"])
        ws.cell(1, 1).value = "Command"
        ws.cell(1, 2).value = "Selector_type"
        ws.cell(1, 3).value = "Selector"
        ws.cell(1, 4).value = "Text or URL"
        ws.cell(1, 5).value = "Expected"
        ws.cell(1, 6).value = "Actual"
        ws.cell(1, 7).value = "Description (50 characters or less for best presentation results)"
        wb.save(self.file_name)

    def check_file_name_unique(self):
        # new_file_Name = self.file_name
        if not os.path.exists(self.file_name):
            return

        print(f"self.file_name = {self.file_name}")
        for x in range(1,100):
            new_file_Name = self.file_name
            new_file_Name = new_file_Name.replace(".xlsx", str(x) + ".xlsx")
            print(f"new_file_Name = {new_file_Name}")
            if not os.path.exists(new_file_Name):
                self.file_name = new_file_Name
                break

    def check_for_one_trust_consent(self):
        selector_type = "css_selector"
        selector = "#onetrust-accept-btn-handler"
        text_url = ""
        expected = ""
        actual = ""
        try:
            element = self.funct.get_element(selector_type, selector)
            if element is not None:
                command = "Click"
                description = "Click Accept One Trust Cookie Consent button"
                self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)
        except Exception as e:
            print(f"Error in check for one trust consent method:\n{e}")

    def create_click_command(self, element):
        command = "Click"
        selector_type = "xpath"
        selector = self.funct.get_element_xpath(element)
        text_url = ""
        expected = ""
        actual = ""
        if element.get_attribute("type") == "checkbox":
            description = "Click checkbox option"
        else:
            description = "Click radio option"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    def check_for_analytics_tagging(self):
        har_content = self.funct.har_file_contents
        if variables.ga4_analytics_identifier in har_content and variables.google_analytics_identifier in har_content:
            # if self.has_save_complete:
            print("GA4 Analytics Tags Found!")
            ar_har_content = har_content.split("\n")
            for har_entry in ar_har_content:
                if variables.ga4_analytics_identifier in har_entry:
                    ar_har_values = urlparse(har_entry)
                    params = parse_qs(ar_har_values.query)
                    event_name = params.get("en", [None])[0]
                    document_location = params.get("dl", [None])[0]
                    hit_type = params.get("t", [None])[0]
                    tracking_id = params.get("tid", [None])[0]
                    # if tracking_id is not None and event_name is not None:
                    if tracking_id is not None:
                        command = "Check Page Tagging"
                        selector_type = "google-analytics"
                        selector = ""
                        text_url = ""
                        expected = f"en={event_name},tid={tracking_id}"
                        actual = ""
                        description = "Check Google Tag"
                        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    def check_for_media_tagging(self):
        har_content = self.funct.har_file_contents
        ar_har_content = har_content.split("\n")
        for platforms in variables.media_tags:
            # print(f"Platform: {platform[0]}")
            platform = platforms[0]
            tag_params = ""
            if platform in har_content:
                for har_entry in ar_har_content:
                    for tag in platforms[1:]:
                        print(f"  Tag: {tag}")
                        ar_har_values = urlparse(har_entry)
                        params = parse_qs(ar_har_values.query)
                        param_name = params.get(tag, [None])[0]
                        if param_name is not None and len(param_name) > 0:
                            tag_params += f"{tag}={param_name},"
                    if len(tag_params) > 0:
                        command = "Check Page Tagging"
                        selector_type = platform
                        selector = ""
                        text_url = ""
                        expected = tag_params.rstrip(',')
                        actual = ""
                        description = f"Check {platform} Tags"
                        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)
                        tag_params = ""

    def create_screenshot_command(self):
        if self.funct.driver is not None and self.funct.driver.current_url is not None:
            file_name = self.funct.driver.current_url.replace("/","-").replace(":","&").replace(".","_") + ".png"
        else:
            file_name = self.file_name.replace("/data/","")
        command = "Take Screenshot"
        selector_type = ""
        selector = ""
        text_url = file_name
        expected = ""
        actual = ""
        description = f"Screenshot of {self.file_name}"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)

    def create_select_command(self, element):
        options = element.find_elements("tag name", "option")
        random_option = random.choice(options)
        command = "Select Dropdown By Value"
        text_url = random_option.get_attribute('value')
        selector_type = "xpath"
        selector = self.funct.get_element_xpath(element)
        expected = random_option.text
        actual = ""
        description = f"Select random dropdown option: ({expected})"
        self.write_command_to_file(command, selector_type, selector, text_url, expected, actual, description)



